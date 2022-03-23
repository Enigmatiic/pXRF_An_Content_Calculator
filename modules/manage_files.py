"""
Module for xls file manager
"""
from openpyxl import load_workbook, Workbook

class manage_file():
    def __init__(self):
        super(manage_file, self).__init__()

    def open_xlsx_file(file_path):
        # Address of the file on the local computer, i.e, path location
        loc_file = (str(file_path))

        # To open Workbook we declare a hadling variable wb
        wb = load_workbook(loc_file)
        sheet = wb.sheetnames
        # Extract all values
        for row in sheet['Brute Data'].values:
            for value in row:
                print(value)
            print('\n')

        # Extract   all Columns of the workBook Sheet



mf = manage_file()

mf.open_xlsx_file('G:/pfe-uqac/Files to use/pXRF/JTSS_PRCS.xlsx')