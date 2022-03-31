"""
Module for xls file manager
"""
from openpyxl import load_workbook, Workbook

from modules import utils


class ManageFile:
    def __init__(self):
        """Init class"""
        super(ManageFile, self).__init__()

    @staticmethod
    def open_xlsx_file_(file_path):
        """
        Open xlsx file
        :param file_path: Location address of the xlsx file
        :return: workbook and first sheet name who contain all datas
        """
        # Address of the file on the local computer, i.e, path location
        loc_file = (str(file_path))
        # To open Workbook we declare a handling variable wb+
        w_book = load_workbook(loc_file)
        w_sheet = w_book.sheetnames
        # Extract all values
        return w_book, w_sheet[0]

    @staticmethod
    def compile_xlsx_to_dict(workbook, sheet_name):
        """
        Compile file into a dictionary
        :param sheet_name: active sheet_name of xlsx workbook file
        :param workbook: xlsx workbook file
        :return: First line of workbook and  a dictionary who contain all values
        """
        data_head = []
        datas = {}
        # Take first line like table Head
        for row in workbook[sheet_name].values:
            for value in row:
                data_head.append(str(value).lower())
            break

        # Compile values in a dictionary
        i = 0
        for row in workbook[sheet_name].values:
            if i == 0:  # Skip first line
                i += 1
                continue
            i += 1
            j = 0
            datas[str(row[0])] = {}  # Create dict for sample key
            for value in row:
                if j == 0:  # Skip first value line
                    j += 1
                    continue
                if type(value).__name__ == 'str':  # Replace non-analyse element by 0
                    datas[str(row[0])][data_head[j]] = 0
                else:
                    datas[str(row[0])][data_head[j]] = value
                j += 1

        return data_head, datas

    @staticmethod
    def compile_xlsx_calibration_to_dict(workbook, sheet_name):
        """
        Compile calibration file into a dictionary
        :param sheet_name: active sheet_name of xlsx workbook file
        :param workbook: xlsx workbook file
        :return: First line of workbook and  a dictionary who contain all values
        """
        data_head, list_sample = [], []
        datas = {}
        line = 0
        _values = workbook[sheet_name].values
        for __row in _values:
            if line == 0:
                line += 1
                continue
            list_sample.append(__row[0])
        frequency_sample = utils.calibration_sample_frequencies(list_sample)
        # Take first line like table Head
        for _row in workbook[sheet_name].values:
            for val in _row:
                data_head.append(str(val).lower())
            break
        # Compile values in a dictionary
        for sample in frequency_sample:
            datas[sample] = {}
            i, sample_id = 0, 1
            for row in workbook[sheet_name].values:
                if i == 0:  # Skip first line
                    i += 1
                    continue
                if sample == str(row[0]):
                    datas[sample][sample_id] = {}
                    j = 1
                    for value in row[1:]:
                        if type(value).__name__ == 'str':  # Replace non-analyse element by 0
                            datas[sample][sample_id][data_head[j]] = 0
                        else:
                            datas[sample][sample_id][data_head[j]] = value
                        j += 1
                    sample_id += 1
                else:
                    continue

        return data_head, datas, frequency_sample

    def create_xlsx_file(self):
        pass
